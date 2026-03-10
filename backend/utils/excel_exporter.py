"""
====================================================
  Excel Exporter
  ─────────────────────────────────────────────────
  Creates a colour-coded Excel workbook that your
  father can open in Microsoft Excel or Google Sheets.

  Columns exported:
    Stock Name | Symbol | Price | Change% | Signal
    Score | Risk | Technical Score | Explanation
====================================================
"""

import asyncio
import logging
import os
import tempfile
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, GradientFill)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Colour palette ────────────────────────────────────────
COLORS = {
    "STRONG BUY":  {"fill": "00695c", "font": "FFFFFF"},   # dark teal
    "BUY":         {"fill": "388e3c", "font": "FFFFFF"},   # green
    "HOLD":        {"fill": "f9a825", "font": "212121"},   # amber
    "SELL":        {"fill": "e53935", "font": "FFFFFF"},   # red
    "STRONG SELL": {"fill": "880e4f", "font": "FFFFFF"},   # dark red
}
HEADER_FILL   = PatternFill("solid", fgColor="1a237e")   # navy
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(bold=True, size=14, color="1a237e")
UP_FILL       = PatternFill("solid", fgColor="e8f5e9")   # light green
DOWN_FILL     = PatternFill("solid", fgColor="ffebee")   # light red

thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


class ExcelExporter:

    async def export(self, stocks: List[dict], market: str = "IN") -> str:
        """Write Excel file to a temp path and return the path."""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, self._write, stocks, market)

    def _write(self, stocks: List[dict], market: str) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trading Signals"

        # ── Title row ─────────────────────────────────────
        now_str = datetime.now().strftime("%d %b %Y  %I:%M %p")
        ws.merge_cells("A1:I1")
        title_cell       = ws["A1"]
        title_cell.value = f"AI Intraday Trading Signals  —  {now_str}  |  Market: {'India 🇮🇳' if market == 'IN' else 'USA 🇺🇸'}"
        title_cell.font  = TITLE_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Disclaimer row
        ws.merge_cells("A2:I2")
        disc       = ws["A2"]
        disc.value = "⚠️  DISCLAIMER: This is for educational purposes only. Trading involves risk. Always consult a financial advisor before investing."
        disc.font  = Font(italic=True, color="DC3545", size=9)
        disc.alignment = Alignment(horizontal="center")

        # ── Header row ────────────────────────────────────
        headers = [
            "Stock Name", "Symbol", "Price (₹/$)", "Change %",
            "Signal", "AI Score", "Risk Level",
            "Technical Score", "Explanation"
        ]
        for col_idx, header in enumerate(headers, 1):
            cell              = ws.cell(row=3, column=col_idx, value=header)
            cell.fill         = HEADER_FILL
            cell.font         = HEADER_FONT
            cell.alignment    = Alignment(horizontal="center", vertical="center")
            cell.border       = BORDER
        ws.row_dimensions[3].height = 22

        # ── Data rows ─────────────────────────────────────
        for row_idx, stock in enumerate(stocks, start=4):
            signal  = stock.get("signal", "HOLD")
            change  = stock.get("change_pct", 0)
            row_bg  = UP_FILL if change >= 0 else DOWN_FILL

            values = [
                stock.get("name", ""),
                stock.get("symbol", ""),
                stock.get("current_price", 0),
                f"{'+' if change >= 0 else ''}{change:.2f}%",
                signal,
                f"{stock.get('score', 50):.1f} / 100",
                stock.get("risk", "MEDIUM"),
                f"{stock.get('tech_score', 50):.1f} / 100",
                stock.get("explanation", ""),
            ]

            for col_idx, value in enumerate(values, 1):
                cell           = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border    = BORDER
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="center",
                    horizontal="center" if col_idx != 9 else "left"
                )

                # Signal cell gets colour
                if col_idx == 5 and signal in COLORS:
                    c    = COLORS[signal]
                    cell.fill = PatternFill("solid", fgColor=c["fill"])
                    cell.font = Font(bold=True, color=c["font"])
                # Change % colour
                elif col_idx == 4:
                    cell.font = Font(
                        bold=True,
                        color="155724" if change >= 0 else "721c24"
                    )
                else:
                    cell.fill = row_bg

            ws.row_dimensions[row_idx].height = 45

        # ── Column widths ─────────────────────────────────
        col_widths = [22, 14, 14, 12, 14, 12, 12, 16, 55]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # ── Summary tab ───────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        from collections import Counter
        counts = Counter(s.get("signal") for s in stocks)
        ws2["A1"] = "Signal Summary"
        ws2["A1"].font = TITLE_FONT
        ws2["A3"] = "Signal"; ws2["B3"] = "Count"
        for row, (sig, cnt) in enumerate(counts.items(), 4):
            ws2.cell(row=row, column=1, value=sig)
            ws2.cell(row=row, column=2, value=cnt)

        # ── Save to temp file ─────────────────────────────
        tmp  = tempfile.NamedTemporaryFile(
            suffix=".xlsx", delete=False,
            dir=tempfile.gettempdir()
        )
        path = tmp.name
        tmp.close()
        wb.save(path)
        logger.info(f"Excel exported to {path}")
        return path
